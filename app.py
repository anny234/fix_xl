import streamlit as st
import os
from bs4 import BeautifulSoup
import os
import re
import io
import sys
import logging
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def process_file(content):
    parsed_rows = []
    temp_items = []
    metadata = {"תעודה": "", "אסמכתא": "", "תאריך": ""}

    soup = BeautifulSoup(content, 'html.parser')
    for row in soup.find_all('tr'):
        
        if 'mso-outline-level:6' in row.get_attribute_list("style"):
            tds = [td.get_text(strip=True) for td in row.find_all('td')]
            item = {
                    'פריט': tds[1],
                    'תיאור פריט': tds[2],
                    'הזמנה': tds[3],
                    'כמות': pd.to_numeric(tds[4], errors='coerce') if tds[4] else 0,
                    'תאור י.מידה': tds[5],
                    'מחיר ספק': pd.to_numeric("0" + tds[6], errors='coerce') if tds[6] else 0.0,
                    'ערך ספק': pd.to_numeric("0" + tds[7], errors='coerce') if tds[7] else 0.0,
                }
            temp_items.append(item)
        elif 'mso-outline-level:5' in row.get_attribute_list("style"):
            row.find_next("td").get_text(strip=True)
            metadata["תעודה"] = row.find_next("td").get_text(strip=True)
        elif 'mso-outline-level:4' in row.get_attribute_list("style"):
            row.find_next("td").get_text(strip=True)
            metadata["אסמכתא"] = row.find_next("td").get_text(strip=True)
        elif 'mso-outline-level:3' in row.get_attribute_list("style"):
            row.find_next("td").get_text(strip=True)
            metadata["תאריך"] = row.find_next("td").get_text(strip=True)
            
            for item in temp_items:
                combined_row = item.copy()
                combined_row.update(metadata)
                parsed_rows.append(combined_row)
            temp_items = []
            metadata = {"תעודה": "", "אסמכתא": "", "תאריך": ""}
    if not parsed_rows:
        raise ValueError("לא נמצאו שורות תואמות לעיבוד בקובץ שנבחר.")
    df = pd.DataFrame(parsed_rows)
    
    # Crucial Step: Parse text date fragments ("DD.MM.YY") into real Timestamp elements
    df['תאריך'] = pd.to_datetime(df['תאריך'], format='%d.%m.%y', errors='coerce')
    
    cols = ['תאריך', 'אסמכתא', 'תעודה', 'פריט', 'תיאור פריט', 'הזמנה', 'כמות', 'תאור י.מידה', 'מחיר ספק', 'ערך ספק']
    df = df[cols]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "דוח ספקים מרוכז"
    
    ws.views.sheetView[0].showGridLines = True
    ws.sheet_view.rightToLeft = True
    
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
        
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=11)
    zebra_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=1):
        is_header = (row_idx == 1)
        for col_idx, cell in enumerate(row, start=1):
            cell.border = thin_border
            if is_header:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.font = data_font
                if row_idx % 2 == 0:
                    cell.fill = zebra_fill
                
                col_name = df.columns[col_idx - 1]
                
                # Configuration updates for data types and alignments
                if col_name == 'תאריך':
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.number_format = 'yyyy-mm-dd' # Formats cell as date for Excel sorting tools
                elif col_name in ['כמות', 'מחיר ספק', 'ערך ספק']:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = '#,##0' if col_name == 'כמות' else '#,##0.00'
                elif col_name in ['פריט', 'אסמכתא', 'תעודה']:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.number_format = '@' # Enforces clean text representation for reference IDs
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(ws.max_column)}{ws.max_row}"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer.read()  # bytes


st.title("Fix XL dates")

# 1. File Upload Component
uploaded_file = st.file_uploader("Choose a file to process")

if uploaded_file is not None:
    # Read the file data
    file_bytes = uploaded_file.read()
    
    st.info("Processing your file...")

    output_data = process_file(file_bytes) 
    output_filename = f"processed_{uploaded_file.name}"
    
    st.success("Done!")
    
    # 3. File Download Component
    st.download_button(
        label="Download Processed File",
        data=output_data,
        file_name=output_filename
    )
